# -*- coding: utf-8 -*-
"""
pmj-keieidangi : 経営談義 AI財務分析 Cloud Run サービス (Flask)

エンドポイント:
  GET  /         : ヘルスチェック
  POST /analyze  : 財務データをAI分析して返す(非ストリーミング)
      body(JSON): { "report": {...報告書JSON...}, "tone": "expert|plain",
                    "providers": ["claude","gemini","openai"] }
      返り値    : { "status":"OK", "tone":..., "providers":[...],
                    "results": { provider: {"sections":{...}} | {"error":...} } }

APIキーは環境変数で渡す(Cloud Runの環境変数/Secret):
  ANTHROPIC_API_KEY, OPENAI_API_KEY, GOOGLE_API_KEY
  任意: ANTHROPIC_MODEL, OPENAI_MODEL, GEMINI_MODEL
"""
from __future__ import annotations
import os, re, threading
from io import BytesIO
from flask import Flask, request, jsonify, send_file

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "keiei_template.xlsm")

# ---- トーン --------------------------------------------------------------
TONE_INSTRUCTIONS = {
    "expert": (
        "【表現: 財務・会計の専門用語を積極的に使用】\n"
        "経理部門・金融機関担当者・税理士などが読む前提で、財務会計の正式な専門用語を正しく使ってください。\n"
        "■推奨用語例: 『売上総利益率(粗利率)』『営業利益率』『売上債権回転日数(DSO)』『棚卸資産回転日数』\n"
        "　『買入債務回転日数』『自己資本比率』『固定長期適合率』『流動比率』『当座比率』『EBITDA』\n"
        "　『キャッシュ・コンバージョン・サイクル(CCC)』『損益分岐点売上高(BEP)』『営業レバレッジ』\n"
        "　『運転資本』『インタレスト・カバレッジ・レシオ』『ROA』『ROE』『販管費率』 等を状況に応じて使用。\n"
        "■数値表現は専門家の書式で。例: 『粗利率14.31%(前期比▲2.73pt)』『DSO 20.8日(+3.1日)』。\n"
        "■省略せずに正確な勘定科目名・指標名を用いる。略語は初出時に日本語名を併記。\n"
        "■文体: 常体(である調)で簡潔に。一般人向けの補足説明は不要。"
    ),
    "plain": (
        "【表現: 財務・会計の専門用語を極力使用しない】\n"
        "財務知識のない中小企業の経営者・店長が読んで直感的に理解できる、平易な日本語で書いてください。\n"
        "■専門用語は原則として使わない。やむを得ず使う場合は、必ず括弧書きで平易な言い換えを添える。\n"
        "　例: 『売掛金（商品を売ったがまだ受け取っていないお金）』『自己資本比率（会社の体力を示す割合）』。\n"
        "■避ける語: 『DSO』『EBITDA』『CCC』『営業レバレッジ』等の略語・カタカナ財務用語は使用禁止。\n"
        "　代わりに『売掛金の回収までの日数』『本業の稼ぐ力』等の日常語に。\n"
        "■数値表現は日常感覚で。例: 『現金が約4日分しか手元にない』『前の年より利益が約1,300万円減った』。\n"
        "■文体: 敬体(です・ます調)でやさしく。難しい表現は避け、1文を短めに。比喩や身近な例えを活用。"
    ),
}

SECTION_NAMES = ["REPORT", "SALES_ISSUE", "SALES_PROPOSAL",
                 "INCOME_ISSUE", "INCOME_PROPOSAL", "CAPITAL_ISSUE", "CAPITAL_PROPOSAL"]
MARKER_RE = re.compile(r"===\s*(" + "|".join(SECTION_NAMES) + r")\s*===")

# ---- プロンプト ----------------------------------------------------------
def _format_financial(data):
    si = (data or {}).get("store_info", {}) or {}
    fin = (data or {}).get("financials", {}) or {}
    h = fin.get("headers", {}) or {}
    yc, yp, yp2 = h.get("year_current"), h.get("year_previous"), h.get("year_previous2")
    lines = [
        "店コード: %s" % si.get("store_code"),
        "店名: %s" % si.get("store_name"),
        "営業所: %s" % si.get("office"),
        "担当者: %s" % si.get("person_in_charge"),
        "",
        "対象年度: %s年 / 前年: %s年 / 前前年: %s年" % (yc, yp, yp2),
        "",
        "【財務項目】",
    ]
    def fmt(v, u):
        if v is None or v == "":
            return "-"
        unit = u or ""
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return "{:,.2f}{}".format(v, unit)
        return "{}{}".format(v, unit)
    for r in fin.get("rows", []):
        lines.append(
            "[{}] {} {}: {}={}, {}={}, 前期差={}, {}={}, 前前期差={}".format(
                r.get("category") or "", r.get("no") or "", r.get("item") or "",
                yc, fmt(r.get("v_curr"), r.get("u_curr")),
                yp, fmt(r.get("v_prev"), r.get("u_prev")),
                fmt(r.get("diff_prev"), r.get("u_diff")),
                yp2, fmt(r.get("v_prev2"), r.get("u_prev2")),
                fmt(r.get("diff_prev2"), r.get("u_diff2")),
            )
        )
    # 全国平均(ベンチマーク): 直近決算の販売高から年商ランクを判定した全国平均。自店と比較する材料
    reg = (data or {}).get("regional") or {}
    gm, ph = reg.get("gross_margin_avg"), reg.get("ph_avg")
    if gm is not None or ph is not None:
        lines.append("")
        lines.append("【全国平均（ベンチマーク）】※直近決算の販売高から判定した年商ランクの全国平均。自店の粗利益率・販売P・Hと比較して講評すること")
        lines.append("年商ランク: %s%s" % (reg.get("rank") or "-",
                                        ("（%s）" % reg.get("span")) if reg.get("span") else ""))
        lines.append("全国平均 粗利益率: %s" % ("%.1f%%" % gm if gm is not None else "-"))
        lines.append("全国平均 販売P・H: %s" % ("{:,.0f}千円".format(ph) if ph is not None else "-"))
    return "\n".join(lines)

def build_prompt(data, tone):
    aliases = {"mild": "plain", "normal": "expert", "strict": "expert"}
    tone = aliases.get(tone, tone)
    tone = tone if tone in TONE_INSTRUCTIONS else "expert"
    body = _format_financial(data)
    return (
        "あなたは中小企業の経営コンサルタントです。下記の財務データを分析し、\n"
        "『経営談義報告書』と、販売・収支・資金それぞれの『課題』と『提案』を日本語で作成してください。\n\n"
        "【表現ルール — 100%遵守すること】\n"
        + TONE_INSTRUCTIONS[tone] + "\n\n"
        "■上記の表現ルールは出力全体（REPORT/各課題/各提案すべて）に適用すること。\n\n"
        "【出力フォーマット (厳守)】\n"
        "必ず下記のマーカー形式で出力してください。余計な前置き・挨拶は不要です。\n"
        "マーカーは半角英大文字のみ。課題と提案は明確に分けて書くこと。\n\n"
        "===REPORT===\n(経営談義報告書として、全体状況・キーメッセージを **300文字以内** で要約)\n\n"
        "===SALES_ISSUE===\n(販売の課題。箇条書き3〜5点。販売高/月商/ペイライン/従業員生産性の観点)\n\n"
        "===SALES_PROPOSAL===\n(販売の提案。課題に対応する具体的な改善アクション。箇条書き3〜5点)\n\n"
        "===INCOME_ISSUE===\n(収支の課題。箇条書き3〜5点。粗利率・管理経費率・営業利益率・支払利息に言及)\n\n"
        "===INCOME_PROPOSAL===\n(収支の提案。課題に対応する改善アクション。箇条書き3〜5点)\n\n"
        "===CAPITAL_ISSUE===\n(資金の課題。箇条書き3〜5点。現預金日数・売掛/棚卸/借入日数・自己資本に言及)\n\n"
        "===CAPITAL_PROPOSAL===\n(資金の提案。課題に対応する改善アクション。箇条書き3〜5点)\n\n"
        "【財務データ】\n" + body + "\n"
    )

def render_analysis_template(template, data, tone):
    """外部プロンプト雛形に {tone}/{fin} を差し込む。{fin} が無ければデータを末尾に補完。"""
    aliases = {"mild": "plain", "normal": "expert", "strict": "expert"}
    tone = aliases.get(tone, tone)
    tone = tone if tone in TONE_INSTRUCTIONS else "expert"
    fin = _format_financial(data)
    p = template.replace("{tone}", TONE_INSTRUCTIONS[tone]).replace("{fin}", fin)
    if "{fin}" not in template:
        p += "\n\n【財務データ】\n" + fin
    return p

def split_sections(text):
    out = {}
    if not text:
        return out
    matches = list(MARKER_RE.finditer(text))
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        out[m.group(1)] = text[start:end].strip()
    if not matches:
        out["REPORT"] = text.strip()
    return out

# ---- 各プロバイダ (Claude のみストリーミング) --------------------------
# 出力トークン上限 / API クライアントのタイムアウト(秒)。
# 要約は最大 28000 文字程度になるため大きめに取る。
MAX_OUTPUT_TOKENS = 32000
CLIENT_TIMEOUT = 900.0

# 優先モデル。使えなければ API のモデル一覧から利用可能なものへ自動フォールバック
PREFERRED_CLAUDE_MODEL = "claude-sonnet-4-6"
_claude_model_cache = None

def _resolve_claude_model(client):
    """claude-sonnet-4-6 が利用可能ならそれを使う。
       使えない場合は API の models.list() から利用可能なモデル(sonnet優先→最新)を選ぶ。
       環境変数 ANTHROPIC_MODEL は使用しない。"""
    global _claude_model_cache
    if _claude_model_cache:
        return _claude_model_cache
    chosen = PREFERRED_CLAUDE_MODEL
    try:
        ids = [m.id for m in client.models.list(limit=1000).data]
        if PREFERRED_CLAUDE_MODEL not in ids:
            sonnets = sorted([m for m in ids if "sonnet" in m], reverse=True)
            chosen = sonnets[0] if sonnets else (ids[0] if ids else PREFERRED_CLAUDE_MODEL)
    except Exception:
        chosen = PREFERRED_CLAUDE_MODEL  # 一覧取得不可時は優先モデルで試行
    _claude_model_cache = chosen
    return chosen

def call_claude(prompt):
    import anthropic
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        raise RuntimeError("ANTHROPIC_API_KEY 未設定")
    client = anthropic.Anthropic(api_key=key, timeout=CLIENT_TIMEOUT)
    model = _resolve_claude_model(client)
    # max_tokens が大きいと SDK が「10分を超える可能性のある非ストリーミング要求」として
    # 送信前に拒否する。要約は長文になるため必ずストリーミングで受け取って連結する。
    parts = []
    with client.messages.stream(
        model=model, max_tokens=MAX_OUTPUT_TOKENS,
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        for t in stream.text_stream:
            parts.append(t)
    return "".join(parts)

# OpenAI 優先モデル。使えなければ models.list() から利用可能な gpt 系へフォールバック
PREFERRED_OPENAI_MODEL = "gpt-4o-mini"
_openai_model_cache = None

def _resolve_openai_model(client):
    """gpt-4o-mini が使えればそれ。無ければ gpt 系チャットモデルの利用可能な最新を選ぶ。
       環境変数 OPENAI_MODEL は使用しない。"""
    global _openai_model_cache
    if _openai_model_cache:
        return _openai_model_cache
    chosen = PREFERRED_OPENAI_MODEL
    try:
        ids = [m.id for m in client.models.list().data]
        if PREFERRED_OPENAI_MODEL not in ids:
            BLOCK = ("embedding", "whisper", "tts", "dall-e", "image", "audio",
                     "realtime", "moderation", "transcribe", "search")
            usable = [i for i in ids if i.startswith("gpt-") and not any(b in i for b in BLOCK)]
            for p in ("gpt-4o-mini", "gpt-4o", "gpt-4.1", "gpt-4", "gpt-"):
                cands = sorted([i for i in usable if i.startswith(p)], reverse=True)
                if cands:
                    chosen = cands[0]
                    break
    except Exception:
        chosen = PREFERRED_OPENAI_MODEL
    _openai_model_cache = chosen
    return chosen

def call_openai(prompt):
    from openai import OpenAI
    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        raise RuntimeError("OPENAI_API_KEY 未設定")
    client = OpenAI(api_key=key, timeout=CLIENT_TIMEOUT)
    model = _resolve_openai_model(client)
    r = client.chat.completions.create(
        model=model, temperature=0.3,
        messages=[{"role": "user", "content": prompt}],
    )
    return r.choices[0].message.content or ""

# Gemini 優先モデル。使えなければ models.list()(generateContent対応)から最新へフォールバック
PREFERRED_GEMINI_MODEL = "gemini-2.5-flash"
_gemini_model_cache = None

def _resolve_gemini_model(client):
    """gemini-2.5-flash が使えればそれ。無ければ generateContent 対応の gemini 系の最新を選ぶ。
       環境変数 GEMINI_MODEL は使用しない。"""
    global _gemini_model_cache
    if _gemini_model_cache:
        return _gemini_model_cache
    chosen = PREFERRED_GEMINI_MODEL
    try:
        names = []
        for m in client.models.list():
            n = (getattr(m, "name", "") or "").split("/")[-1]
            if not n:
                continue
            acts = getattr(m, "supported_actions", None) or getattr(m, "supported_generation_methods", None) or []
            if acts and not any("generateContent" in str(a) for a in acts):
                continue
            names.append(n)
        if PREFERRED_GEMINI_MODEL not in names:
            for p in ("gemini-2.5-flash", "gemini-2.5", "gemini-flash", "gemini-"):
                cands = sorted([n for n in names if n.startswith(p) and "embedding" not in n], reverse=True)
                if cands:
                    chosen = cands[0]
                    break
    except Exception:
        chosen = PREFERRED_GEMINI_MODEL
    _gemini_model_cache = chosen
    return chosen

def call_gemini(prompt):
    from google import genai
    key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY")
    if not key:
        raise RuntimeError("GOOGLE_API_KEY 未設定")
    client = genai.Client(api_key=key)
    model = _resolve_gemini_model(client)
    # 要約を長く出すため出力上限を明示(未対応SDKでは指定なしで実行)
    cfg = None
    try:
        from google.genai import types
        cfg = types.GenerateContentConfig(max_output_tokens=MAX_OUTPUT_TOKENS)
    except Exception:
        cfg = None
    if cfg is not None:
        r = client.models.generate_content(model=model, contents=prompt, config=cfg)
    else:
        r = client.models.generate_content(model=model, contents=prompt)
    return getattr(r, "text", "") or ""

PROVIDERS = {"claude": call_claude, "openai": call_openai, "gemini": call_gemini}

def _run(prov, prompt, results):
    try:
        results[prov] = {"sections": split_sections(PROVIDERS[prov](prompt))}
    except Exception as e:
        results[prov] = {"error": str(e)[:300]}

def analyze(data, tone, providers, template=None):
    # 外部プロンプト雛形(zaiTask の analysisprompt)があればそれを使用。無ければ内蔵。
    if isinstance(template, str) and template.strip():
        prompt = render_analysis_template(template, data, tone)
    else:
        prompt = build_prompt(data, tone)
    results = {}
    ts = []
    for p in providers:
        if p in PROVIDERS:
            t = threading.Thread(target=_run, args=(p, prompt, results), daemon=True)
            t.start(); ts.append(t)
    for t in ts:
        t.join(timeout=180)
    return results

# ---- ルーティング --------------------------------------------------------
@app.get("/")
def health():
    return jsonify({"status": "ok", "service": "pmj-keieidangi"})

@app.post("/analyze")
def analyze_route():
    body = request.get_json(silent=True) or {}
    data = body.get("report") or {}
    tone = body.get("tone", "expert")
    if tone not in TONE_INSTRUCTIONS:
        tone = "expert"
    raw = body.get("providers")
    if isinstance(raw, str):
        raw = [p.strip() for p in raw.split(",")]
    providers = [p for p in (raw or ["claude", "openai", "gemini"]) if p in PROVIDERS]
    if not providers:
        providers = ["claude", "openai", "gemini"]
    if not data.get("financials"):
        return jsonify({"status": "NG", "error": "report.financials がありません"}), 400
    results = analyze(data, tone, providers, body.get("prompt_template"))
    return jsonify({"status": "OK", "tone": tone, "providers": providers, "results": results})

# ---- AI自動要約 (複数AIの分析を比較・検証) -----------------------------
PROVIDER_LABEL = {"claude": "Claude", "gemini": "Gemini", "openai": "OpenAI"}
SEC_LABELS = {
    "REPORT": "経営談義報告書",
    "SALES_ISSUE": "販売-課題", "SALES_PROPOSAL": "販売-提案",
    "INCOME_ISSUE": "収支-課題", "INCOME_PROPOSAL": "収支-提案",
    "CAPITAL_ISSUE": "資金-課題", "CAPITAL_PROPOSAL": "資金-提案",
}
SUMMARY_ORDER = ["claude", "openai", "gemini"]  # 要約担当の優先順

def _has_key(p):
    if p == "claude":
        return bool(os.environ.get("ANTHROPIC_API_KEY"))
    if p == "openai":
        return bool(os.environ.get("OPENAI_API_KEY"))
    if p == "gemini":
        return bool(os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY"))
    return False

def _format_results(results):
    """各AIの分析テキストを読みやすいブロックに整形(存在するAIのみ)。"""
    blocks = []
    for p in SUMMARY_ORDER:
        r = (results or {}).get(p)
        if not r:
            continue
        if r.get("error"):
            blocks.append("=== %s の分析: エラー(%s) ===" % (PROVIDER_LABEL.get(p, p), r.get("error")))
            continue
        sec = r.get("sections") or {}
        lines = ["=== %s の分析 ===" % PROVIDER_LABEL.get(p, p)]
        for key in SECTION_NAMES:
            v = sec.get(key)
            if v:
                lines.append("【%s】\n%s" % (SEC_LABELS.get(key, key), v))
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks)

# 要約も分析側と同じ「課題/提案」区分で出力させる(SUM_*_ISSUE / SUM_*_PROPOSAL)。
# 旧形式(SUM_SALES など課題・提案が一体)のマーカーも受理し、画面/Excelが空にならないようにする。
SUMMARY_CATS = ["SALES", "INCOME", "CAPITAL"]
SUMMARY_SECTIONS = (["SUM_REPORT"]
                    + ["SUM_%s_%s" % (c, s) for c in SUMMARY_CATS for s in ("ISSUE", "PROPOSAL")]
                    + ["SUM_%s" % c for c in SUMMARY_CATS])   # 末尾3つは旧形式(後方互換)
SUMMARY_MARKER_RE = re.compile(r"===\s*(" + "|".join(SUMMARY_SECTIONS) + r")\s*===")
SUMMARY_KEYMAP = dict((s, s[4:]) for s in SUMMARY_SECTIONS)   # "SUM_" を落としたものがキー

def split_summary(text):
    out = dict((k, "") for k in SUMMARY_KEYMAP.values())
    out["REPORT"] = ""
    if not text:
        return out
    matches = list(SUMMARY_MARKER_RE.finditer(text))
    if not matches:
        out["REPORT"] = text.strip()
        return out
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        out[SUMMARY_KEYMAP[m.group(1)]] = text[start:end].strip()
    # 旧形式(課題・提案が一体)しか無い区分は、課題側に寄せて表示できるようにする
    for c in SUMMARY_CATS:
        if out.get(c) and not out.get(c + "_ISSUE") and not out.get(c + "_PROPOSAL"):
            out[c + "_ISSUE"] = out[c]
    return out

# 内蔵の要約プロンプト。外部ファイル(zaiTask の summaryprompt)が無い場合に使う。
# ※ 資材の kr2pmrpcm3pbh6rd655hprompt/summaryprompt と同一内容を保つこと。
SUMMARY_PROMPT_TEMPLATE = """\
あなたは複数の生成AIが作成した経営分析をレビューする上級経営コンサルタントです。
同じ財務データに対して各AIが作成した『経営分析』を比較・統合し、
経営談義の場でそのまま配布・議論できる詳細な『検討資料』を、販売・収支・資金それぞれの
『課題』と『提案』に分けて日本語で作成してください。
（分析が存在するAIのみを対象。AIは1〜3個のいずれの場合もある）

【表現ルール — 100%遵守すること】
{tone}

【分量 — 最重要。必ず守ること】
・要約と称して短くしないこと。各AIの記述を圧縮するのではなく、統合したうえで大幅に掘り下げること。
・===SUM_REPORT=== は 2400〜3600文字。7〜10段落に分けて記述する。
・各『課題』『提案』は箇条書きで12〜18点。1点あたり200〜350文字の小段落として記述する。
・『課題』の各点には次の3要素を必ず含める。
    (1) 該当する数値（金額・比率・日数・前期比・全国平均との差）
    (2) その数値が事業に与える影響
    (3) そうなっている原因の仮説
・『提案』の各点には次の4要素を必ず含める。
    (1) 具体的な施策の内容
    (2) 実行手順（誰が・何を・どの順番で）
    (3) 到達目標の数値
    (4) 実施期限の目安
・各AIに共通する指摘は1点に統合し、1つのAIしか触れていない指摘も重要であれば必ず拾うこと。
・各AIの指摘を拾い切っても点数が足りない場合は、財務データから読み取れる論点を自分で追加すること。
  （例：3期の推移、全国平均との乖離、回転日数のバランス、損益分岐点、資金繰りの余力 など）
・AI間で結論が食い違う点は、その区分の末尾に「※AI間の見解差：…」として2〜3点で明記する。
・「特になし」「同上」「前述のとおり」などで省略しないこと。全区分を最後まで書き切ること。
・同じ内容を言い換えただけの水増しはしないこと。各点は必ず異なる論点を扱うこと。

【出力フォーマット (厳守)】
必ず下記7マーカー形式で出力。余計な前置きは不要。マーカーは半角英大文字のみ。
課題と提案は明確に分けて書くこと。
※ マーカー行(===SUM_REPORT=== / ===SUM_SALES_ISSUE=== / ===SUM_SALES_PROPOSAL=== / ===SUM_INCOME_ISSUE=== / ===SUM_INCOME_PROPOSAL=== / ===SUM_CAPITAL_ISSUE=== / ===SUM_CAPITAL_PROPOSAL===)は変更・削除しないでください。出力の解析に使用します。

===SUM_REPORT===
(全体状況のまとめ。各AIの見解を統合し、3期の業績推移・キーメッセージ・重要な数値・
 全国平均との比較・資金繰りの見通し・AI間の見解差を含めて2400〜3600文字。7〜10段落に分ける)

===SUM_SALES_ISSUE===
(販売の課題を統合。箇条書き12〜18点。販売高/平均月商/ペイライン/従業員生産性の観点)

===SUM_SALES_PROPOSAL===
(販売の提案を統合。課題に対応する具体的な改善アクション。箇条書き12〜18点)

===SUM_INCOME_ISSUE===
(収支の課題を統合。箇条書き12〜18点。粗利率・管理経費率・営業利益率・支払利息に言及)

===SUM_INCOME_PROPOSAL===
(収支の提案を統合。課題に対応する改善アクション。箇条書き12〜18点)

===SUM_CAPITAL_ISSUE===
(資金の課題を統合。箇条書き12〜18点。現預金日数・売掛/棚卸/借入日数・自己資本に言及)

===SUM_CAPITAL_PROPOSAL===
(資金の提案を統合。課題に対応する改善アクション。箇条書き12〜18点)

【財務データ】
{fin}

【各AIの分析】
{analyses}
"""

def build_summary_prompt(data, results, tone):
    return render_summary_template(SUMMARY_PROMPT_TEMPLATE, data, results, tone)

def render_summary_template(template, data, results, tone):
    """外部から渡されたプロンプト雛形に {tone}/{fin}/{analyses} を差し込む。
       雛形に該当プレースホルダが無い場合はデータを末尾に補完(取りこぼし防止)。"""
    aliases = {"mild": "plain", "normal": "expert", "strict": "expert"}
    tone = aliases.get(tone, tone)
    tone = tone if tone in TONE_INSTRUCTIONS else "expert"
    fin = _format_financial(data)
    analyses = _format_results(results)
    p = (template.replace("{tone}", TONE_INSTRUCTIONS[tone])
                 .replace("{fin}", fin)
                 .replace("{analyses}", analyses))
    if "{fin}" not in template:
        p += "\n\n【財務データ】\n" + fin
    if "{analyses}" not in template:
        p += "\n\n【各AIの分析】\n" + analyses
    return p

@app.post("/summarize")
def summarize_route():
    body = request.get_json(silent=True) or {}
    data = body.get("report") or {}
    results = body.get("results") or {}
    tone = body.get("tone", "expert")
    if tone not in TONE_INSTRUCTIONS:
        tone = "expert"
    if not results:
        return jsonify({"status": "NG", "error": "results(各AIの分析)がありません"}), 400

    # 要約担当AIを選定: body.provider 指定 → SUMMARY_PROVIDER 環境変数 → 既定順
    req_prov = (str(body.get("provider") or "")).strip().lower()
    pref = req_prov if req_prov in PROVIDERS else (os.environ.get("SUMMARY_PROVIDER", "") or "").strip().lower()
    order = ([pref] if pref in PROVIDERS else []) + [p for p in SUMMARY_ORDER if p != pref]
    chosen = None
    for p in order:
        if _has_key(p):
            chosen = p
            break
    if not chosen:
        return jsonify({"status": "NG", "error": "要約用のAPIキーが未設定です"}), 200

    # 外部プロンプト雛形(zaiTask の summaryprompt)があればそれを使用。無ければ内蔵。
    template = body.get("prompt_template")
    if isinstance(template, str) and template.strip():
        prompt = render_summary_template(template, data, results, tone)
    else:
        prompt = build_summary_prompt(data, results, tone)
    try:
        text = PROVIDERS[chosen](prompt)
    except Exception as e:
        return jsonify({"status": "NG", "error": str(e)[:300], "provider": chosen}), 200
    return jsonify({"status": "OK", "provider": chosen, "sections": split_summary(text)})


# ---- Excel 生成 (報告書 に入力セルのみ書き込み・数式/地区平均は自動) ---------
# 26年最新版雛形: シート名は「報告書」。対象従業員(12)は貼り付けデータからXLOOKUP、
# 決算書BS(38)は数式で自動化されたため入力行から除外。地区平均は店コード(D4)から自動。
EXCEL_SHEET = "報告書"
# テンプレートの「入力セル」がある行(数式行・自動計算行は書き込まない)
EXCEL_INPUT_ROWS = {9, 15, 17, 19, 21, 23, 25, 27, 29, 31, 33, 35, 37}

# ---- AI診断シート (3AI分析を画面の配色で貼る。A4印刷・1〜3個対応・編集可) ----
AI_SHEET_PROVS = ["claude", "gemini", "openai"]
AI_SHEET_HDR = {"claude": ("Claude", "D77655"), "gemini": ("Gemini", "1A73E8"), "openai": ("OpenAI", "10A37F")}
AI_SHEET_CATS = [("SALES", "販売"), ("INCOME", "収支"), ("CAPITAL", "資金")]

def _est_lines(text, width):
    """セル幅(width)に対するおおよその行数(日本語は全角=2幅でカウント)。"""
    if not text:
        return 1
    total = 0
    for line in str(text).split("\n"):
        dl = 0
        for ch in line:
            dl += 2 if ord(ch) > 0x2E80 else 1
        total += max(1, -(-dl // max(1, int(width))))
    return max(1, total)

def fix_report_print_area(wb, ws):
    """報告書シートの印刷範囲を復元する。

    雛形の Print_Area は INDIRECT(報告書!$X$1) という動的定義になっており、
    X1 の文字列("報告書" / "経営談義資料")が名前定義を指すことで
    印刷範囲を切り替える仕組みになっている。
    openpyxl は INDIRECT を保持できず Print_Area を "$X$1" に潰してしまい、
    印刷すると X1 の1セルしか出力されない。
    そのため生成時に X1 を解決し、実際の範囲を静的に設定し直す。
    """
    ref = None
    try:
        key = ws["X1"].value
        if key:
            key = str(key).strip()
            dn = None
            try:
                dn = ws.defined_names.get(key)      # シートローカル名
            except Exception:
                dn = None
            if dn is None:
                try:
                    dn = wb.defined_names.get(key)  # ブックスコープ名
                except Exception:
                    dn = None
            if dn is not None:
                ref = getattr(dn, "value", None) or getattr(dn, "attr_text", None)
    except Exception:
        ref = None
    if not ref:
        ref = "$A$3:$X$55"                          # フォールバック(雛形の既定範囲)
    if "!" in ref:
        ref = ref.split("!")[-1]
    ws.print_area = ref

def setup_avg_sheet_print(wb):
    """平均粗利益率・PH シートを A4 横・幅1ページ に収める設定にする。
       雛形は縦向き(Zoom=100)のままなので、生成時に上書きする。"""
    name = "平均粗利益率・PH"
    if name not in wb.sheetnames:
        return
    from openpyxl.worksheet.properties import PageSetupProperties
    ws = wb[name]
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.scale = None   # 拡大縮小率をクリア(fitToPage を有効にするため)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

def build_ai_sheet(wb, results, layout="landscape"):
    """AI診断シートに各AIの分析(報告書/販売/収支/資金の課題・提案)を貼る。
       layout: "landscape"(横3列並列) / "portrait"(縦・各AIを全幅で縦積み・AIごと改ページ)
       results が 1〜3個でも 0個でもエラーにしない。要約は貼らない。"""
    results = results or {}
    provs = [p for p in AI_SHEET_PROVS
             if isinstance(results.get(p), dict) and results[p].get("sections")]
    if not provs:
        return  # AI分析結果が無ければシートは触らない

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties
    from openpyxl.worksheet.pagebreak import Break

    FONT = "Yu Gothic"
    def _fill(c):
        return PatternFill("solid", fgColor=c)
    def _font(color="2B2F33", bold=False, size=10, white=False):
        return Font(name=FONT, size=size, bold=bold, color=("FFFFFF" if white else color))
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["AI診断"] if "AI診断" in wb.sheetnames else wb.create_sheet("AI診断")

    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # ============ 縦向き・縦積み(各AIを全幅、AIごとに改ページ) ============
    if layout == "portrait":
        ws.page_setup.orientation = "portrait"
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5
        COLW = 95
        COL = 2  # B列(全幅)
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions[get_column_letter(COL)].width = COLW
        st = {"r": 1}

        def put(text, bg, fg, white=False, bold=True, size=10, wrap=False, h=18, center=False):
            r = st["r"]
            c = ws.cell(r, COL, text)
            c.fill = _fill(bg)
            c.font = _font(color=(fg or "2B2F33"), bold=bold, size=size, white=white)
            if wrap:
                c.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[r].height = min(680, _est_lines(text, COLW) * 15 + 6)
            else:
                c.alignment = Alignment(vertical="center", horizontal=("center" if center else "left"))
                ws.row_dimensions[r].height = h
            c.border = border
            st["r"] = r + 1

        first = True
        for p in provs:
            if not first:
                ws.row_dimensions[st["r"]].height = 8  # 余白
                st["r"] += 1
                ws.row_breaks.append(Break(id=st["r"] - 1))  # 新AIはページ先頭から
            first = False
            name, clr = AI_SHEET_HDR[p]
            sec = results[p].get("sections") or {}
            put(name, clr, None, white=True, bold=True, size=14, h=26, center=True)
            put("経営談義報告書", "FFF8E1", "8D6E00", bold=True, size=10, h=16)
            put(sec.get("REPORT", "") or "", "FFF8E1", None, bold=False, wrap=True)
            for key, title in AI_SHEET_CATS:
                put(title, "0A66C2", None, white=True, bold=True, size=11, h=20)
                put("課題", "FDF6F6", "B42318", bold=True, size=10, h=15)
                put(sec.get(key + "_ISSUE", "") or "", "FDF6F6", None, bold=False, wrap=True)
                put("提案", "F2FBF3", "1A7F37", bold=True, size=10, h=15)
                put(sec.get(key + "_PROPOSAL", "") or "", "F2FBF3", None, bold=False, wrap=True)
        ws.print_area = "A1:%s%d" % (get_column_letter(COL), st["r"] - 1)
        return

    # ============ 横向き・3列並列 ============
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    COLW = 46
    start_col = 2
    ws.column_dimensions["A"].width = 1.5
    for i, p in enumerate(provs):
        ws.column_dimensions[get_column_letter(start_col + i)].width = COLW

    state = {"r": 1}

    def label_row(text, bg, fg, white=False, h=18):
        r = state["r"]
        for i, p in enumerate(provs):
            c = ws.cell(r, start_col + i, text)
            c.fill = _fill(bg)
            c.font = _font(color=(fg or "2B2F33"), bold=True, size=10, white=white)
            c.alignment = Alignment(vertical="center")
            c.border = border
        ws.row_dimensions[r].height = h
        state["r"] = r + 1

    def text_row(section_key, bg):
        r = state["r"]
        maxlines = 1
        for i, p in enumerate(provs):
            txt = (results[p].get("sections") or {}).get(section_key, "") or ""
            c = ws.cell(r, start_col + i, txt)
            c.fill = _fill(bg)
            c.font = _font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
            maxlines = max(maxlines, _est_lines(txt, COLW))
        ws.row_dimensions[r].height = min(640, maxlines * 15 + 6)
        state["r"] = r + 1

    r = state["r"]
    for i, p in enumerate(provs):
        name, clr = AI_SHEET_HDR[p]
        c = ws.cell(r, start_col + i, name)
        c.fill = _fill(clr)
        c.font = _font(white=True, bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[r].height = 22
    state["r"] = r + 1

    label_row("経営談義報告書", "FFF8E1", "8D6E00", h=16)
    text_row("REPORT", "FFF8E1")
    for key, title in AI_SHEET_CATS:
        label_row(title, "0A66C2", None, white=True, h=18)
        label_row("課題", "FDF6F6", "B42318", h=15)
        text_row(key + "_ISSUE", "FDF6F6")
        label_row("提案", "F2FBF3", "1A7F37", h=15)
        text_row(key + "_PROPOSAL", "F2FBF3")

    last_col = get_column_letter(start_col + len(provs) - 1)
    ws.print_area = "A1:%s%d" % (last_col, state["r"] - 1)

# ---- AI診断要約シート (AI自動要約を AI診断 の隣に。画面配色を踏襲) -------
SUMMARY_SHEET_CATS = [("SALES", "販売"), ("INCOME", "収支"), ("CAPITAL", "資金")]

def build_summary_sheet(wb, summary):
    """AI自動要約(報告書/販売/収支/資金 の各要約)を「AI診断要約」シートに貼る。
       summary = {REPORT, SALES, INCOME, CAPITAL}。内容が無ければ作らない。
       配色は画面の AI自動要約パネル準拠(ヘッダ紫/報告書クリーム/カテゴリ青)。"""
    summary = summary or {}
    keys = (["REPORT"]
            + [c + s for c in SUMMARY_CATS for s in ("_ISSUE", "_PROPOSAL")]
            + SUMMARY_CATS)
    if not any(str(summary.get(k) or "").strip() for k in keys):
        return  # 要約が無ければシートは作らない

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    FONT = "Yu Gothic"
    def _fill(c):
        return PatternFill("solid", fgColor=c)
    def _font(color="2B2F33", bold=False, size=10, white=False):
        return Font(name=FONT, size=size, bold=bold, color=("FFFFFF" if white else color))
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    name = "AI診断要約"
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        idx = (wb.sheetnames.index("AI診断") + 1) if "AI診断" in wb.sheetnames else len(wb.sheetnames)
        ws = wb.create_sheet(name, idx)  # AI診断 の隣(右)に作る

    # A4・縦向き・幅1ページに収める(1列・全幅)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5

    COLW = 95
    COL = 2  # B列(全幅)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[get_column_letter(COL)].width = COLW
    st = {"r": 1}

    def put(text, bg, fg, white=False, bold=True, size=10, wrap=False, h=18, center=False):
        r = st["r"]
        c = ws.cell(r, COL, text)
        c.fill = _fill(bg)
        c.font = _font(color=(fg or "2B2F33"), bold=bold, size=size, white=white)
        if wrap:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = min(680, _est_lines(text, COLW) * 15 + 6)
        else:
            c.alignment = Alignment(vertical="center", horizontal=("center" if center else "left"))
            ws.row_dimensions[r].height = h
        c.border = border
        st["r"] = r + 1

    # Excel の行高は最大 409pt。長文を1セルに入れると末尾が印刷されないため、
    # 一定の行数で区切って複数セルに分けて出す(見た目は同じ色で連続する)。
    MAX_LINES = 24

    def put_text(text, bg):
        text = str(text or "")
        if not text.strip():
            put("", bg, None, bold=False, wrap=True)
            return
        chunk, used = [], 0
        for line in text.split("\n"):
            n = _est_lines(line, COLW)
            if chunk and used + n > MAX_LINES:
                put("\n".join(chunk), bg, None, bold=False, wrap=True)
                chunk, used = [], 0
            chunk.append(line)
            used += n
        if chunk:
            put("\n".join(chunk), bg, None, bold=False, wrap=True)

    put("AI自動要約", "6A4EA3", None, white=True, bold=True, size=13, h=24, center=True)  # 紫
    put("経営談義報告書（要約）", "FFF8E1", "8D6E00", bold=True, size=10, h=16)
    put_text(summary.get("REPORT", ""), "FFF8E1")
    for key, title in SUMMARY_SHEET_CATS:
        put(title + "（要約）", "0A66C2", None, white=True, bold=True, size=11, h=20)
        issue = str(summary.get(key + "_ISSUE", "") or "")
        prop = str(summary.get(key + "_PROPOSAL", "") or "")
        if not issue and not prop:
            # 旧形式(課題・提案が一体)の要約はそのまま1ブロックで出す
            put_text(summary.get(key, ""), "FFFFFF")
            continue
        put("課題", "FDF6F6", "B42318", bold=True, size=10, h=15)
        put_text(issue, "FDF6F6")
        put("提案", "F2FBF3", "1A7F37", bold=True, size=10, h=15)
        put_text(prop, "F2FBF3")

    ws.print_area = "A1:%s%d" % (get_column_letter(COL), st["r"] - 1)

@app.post("/excel")
def excel_route():
    body = request.get_json(silent=True) or {}
    data = body.get("report") or {}
    si = data.get("store_info") or {}
    fin = data.get("financials") or {}
    h = fin.get("headers") or {}
    if not fin.get("rows"):
        return jsonify({"status": "NG", "error": "report.financials.rows がありません"}), 400

    import openpyxl
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False, keep_vba=True)
    ws = wb[EXCEL_SHEET]

    # ヘッダー
    sc = si.get("store_code")
    if sc is not None and str(sc).isdigit():
        sc = int(sc)
    ws["D4"] = sc
    ws["D5"] = si.get("store_name")
    ws["X5"] = si.get("office")
    ws["X6"] = si.get("person_in_charge")
    # 決算年
    ws["E7"] = h.get("year_current")
    ws["G7"] = h.get("year_previous")
    ws["K7"] = h.get("year_previous2")
    # 参考値(全国平均): 直近決算の販売高から判定した年商ランクの全国平均を書き込む。
    # 右側(T列)の ○/×/▲ 判定は雛形の数式が自動で行う。単位に注意:
    #   T11: IF(E11<P11,…)        → P11 は千円のまま
    #   T16: IF(E16<P16*100,…)    → P16 は比率(=%÷100)で入れる
    reg = (data.get("regional") or {})
    if reg.get("ph_avg") is not None:
        ws["P11"] = reg.get("ph_avg")
    if reg.get("gross_margin_avg") is not None:
        ws["P16"] = reg.get("gross_margin_avg") / 100.0
    # 入力セルのみ(数式行はスキップ→Excelで再計算)
    for r in fin.get("rows", []):
        rn = r.get("row")
        if rn not in EXCEL_INPUT_ROWS:
            continue
        if r.get("v_curr") not in (None, ""):
            ws["E%d" % rn] = r.get("v_curr")
        if r.get("v_prev") not in (None, ""):
            ws["G%d" % rn] = r.get("v_prev")
        if r.get("v_prev2") not in (None, ""):
            ws["K%d" % rn] = r.get("v_prev2")

    # AI診断シート(3AI分析を貼る。1〜3個でもエラーにしない。要約は貼らない)
    layout = body.get("layout")
    if layout not in ("portrait", "landscape"):
        layout = "portrait"  # 既定: 縦向き・縦積み
    build_ai_sheet(wb, body.get("results"), layout)

    # AI診断要約シート(AI自動要約を AI診断 の隣に。内容が無ければ作らない)
    build_summary_sheet(wb, body.get("summary"))

    # 印刷範囲の復元(雛形の INDIRECT 定義を openpyxl が保持できないため)
    fix_report_print_area(wb, ws)
    setup_avg_sheet_print(wb)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    # ファイル名(Disposition)は呼出側(.do)で付与する想定。ここは本文のみ返す。
    return send_file(
        bio,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
        as_attachment=False,
        download_name="keieidangi.xlsm",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "8080")))
